from openpyxl import Workbook

wb = Workbook()
# ws = wb.active
ws = wb.create_sheet('diary', 0)

data = [('홍길동', 80, 70, 90), ('이기자', 90, 60, 80), ('강기자', 80, 80, 70)]
r = 1
c = 1
for irum, kor, eng, math in data:
    ws.cell(row=r, column=c).value = irum
    ws.cell(row=r, column=c+1).value = kor
    ws.cell(row=r, column=c+2).value = eng
    ws.cell(row=r, column=c+3).value = math
    r += 1

ws.cell(row=r, column=1).value = '합계'
ws.cell(row=r, column=2).value = '=sum(B1:B3)'
ws.cell(row=r, column=3).value = '=sum(C1:C3)'
ws.cell(row=r, column=4).value = '=sum(D1:D3)'

wb.save('openpyxl2.xlsx')
wb.close()